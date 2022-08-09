#imports
import openpyxl
from pathlib import Path

#Data stripped from the excel files
list_of_params = []
list_of_resources = []
list_of_connections = []

#A common parameter between an experiment and a resource
class Connection:
    def __init__(self, rrid, alias, value, units):
        self.alias = alias
        self.rrid = rrid
        self.value = value
        self.units = units

#Equipment used by the researcher
class Resource:
    def __init__(self, RRID, vendor, version, parameter):
        self.RRID = RRID
        self.vendor = vendor
        self.version = version
        self.parameter = parameter

#Environmental conditions to replicate the experiment
class Parameter:
    def __init__(self, alias, value, units):
        self.alias = alias
        self.value = value
        self.units = units


#Script to connect experimental parameters to their respective resources
if __name__ == "__main__":

    #Opening and loading the resources from the experiment
    resource_file = Path('SDS2.0Metadata','resources.xlsx')
    resource_obj = openpyxl.load_workbook(resource_file)

    #Opening and loading the code parameters from the experiment
    param_file = Path('SDS2.0Metadata', 'code_parameters.xlsx' )
    param_obj = openpyxl.load_workbook(param_file)

    param_sheet = param_obj.active
    resource_sheet = resource_obj.active

    #Stripping the resource data
    for i in range(resource_sheet.max_row):
        id = "A" + str(i+1)
        vendor = "C" + str(i+1)
        version = "D" + str(i+1)
        index = "F" + str(i+1)
        list_of_resources.append(Resource(resource_sheet[id].value, resource_sheet[vendor].value,
        resource_sheet[version].value,resource_sheet[index].value))

    #Stripping the parameter data
    for i in range(param_sheet.max_row):
        index = "B" + str(i+1)
        value = "I" + str(i+1)
        units = "H" + str(i+1)
        list_of_params.append(Parameter(param_sheet[index].value, param_sheet[value].value, param_sheet[units].value))

    #Taking out titles
    list_of_resources.pop(0)
    list_of_params.pop(0)
    list_of_params.pop(0)

    #Checking for connections between the resources and experimental parameters
    for x in list_of_resources:
        for p in list_of_params:
            if (x.parameter == p.alias):
                list_of_connections.append(Connection(p.alias,x.RRID, p.value, p.units ))
    
    #Print out connections as proof of concept
    for y in list_of_connections:
        print("Run " + str(y.alias) + " at a " + str(y.rrid) + " of " + str(y.value) + " " + str(y.units))

